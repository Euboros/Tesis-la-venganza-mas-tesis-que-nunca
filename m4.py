#m3v4 3/marzo/2019
#Gabriel Pinto 
#Hernan Campos
import cv2
import numpy as np
import math
from time import time
from openpyxl import Workbook
import os
import datetime


imgi = cv2.imread('izq.jpg')
imgd = cv2.imread('der.jpg')
print("termino carga de fotografias")
tiempo_inicial = time() 
alto, largo,chanels = imgi.shape

#Hacemos un roi a ambas imagenes
c=400
ejey=alto/2-c/2
ejeyy=alto/2+c/2
ejex=largo/2-c/2
ejexx=largo/2+c/2
imgi = imgi[ejey:ejeyy,ejex:ejexx]
imgd = imgd[ejey:ejeyy,ejex:ejexx]
#terminamos roi, tenemos 2 imagenes de 400x400
alto, largo,chanels = imgi.shape
#tomamos las 2 imagenes y eliminamos todo el color azul
print("eliminando azul i")
imgi[:, :, 0] = 0
print("eliminando azul d")
imgd[:, :, 0] = 0

#pasamos imagenes a escala de grises
imgi = cv2.cvtColor(imgi, cv2.COLOR_BGR2GRAY)
imgd = cv2.cvtColor(imgd, cv2.COLOR_BGR2GRAY)
#tenemos 2 imagenes de 400x400 en escala de grises

print("entrando al for de mascara d")
#mascara derecha
for i in range (0,alto):
	for j in range (0,largo):
		if (imgd[i][j]>=30) :
			imgd[i][j]=255
		else:
			imgd[i][j]=0
#mascara izquierda
print("entrando al for de mascara i")
for i in range (0,alto):
	for j in range (largo-1,-1,-1):
		if (imgi[i][j]>=30) :
			imgi[i][j]=255
		else:
			imgi[i][j]=0
pxder = 0 #contador de px negros derecha

pxizq = 0 #izquierda de px negros izquierda

rd  = np.zeros(shape=(12000001), dtype=float)#array para almacenar rugosidad de cada fila derecha
ri  = np.zeros(shape=(12000001), dtype=float)#array para almacenar rugosidad de cada fila izquierda
cod = 0 #contador de cuerpos derecha
coi = 0 #ontador de cuerpos izquierda
h  = 0.0 #altura cuerpo
flagd = True
flagi = True
conversion=13.4#tamano real fotografia(en cm) dividido por los pixeles, ejemplo:una escena de 80.5mm se divide en 6000px siendo estos el largo de la foto
angulo=1.25#tangente
print("calculando lado izq")

for i in range (0,alto):
    #print("fila: " + str(i)) # imprime numero fila
    for j in range (0,largo):
        #izquierda
        if(imgi[i][j] == 0): #si encuentra sombra, cuenta px negros
            flagi = True
            pxizq = pxizq + 1
        else: #si encuentra cuerpo
            if(flagi): #calcula 1 vez por seccion
                ri[coi] =(conversion*pxizq*angulo)#trigonometria para altura acumulada
                coi = coi+1 #aumenta el indice de ri[]
                pxizq = 0 #reinicia cuenta
            flagi = False
    #reinicio de contadores para siguiente iteracion de i
    pxizq = 0
#print ("coi: " + str(coi))
print("calculando lado der")

for i in range (0,alto):
    for j in range (largo-1, -1 , -1):
        #derecha
        if(imgd[i][j] == 0 and j!=0):
            pxder = pxder + 1
            flagd = True
        else:
            if (flagd):
                rd[cod] = (conversion * pxder * angulo)  #trigonometria para altura acumulada
                cod = cod+1
                pxder = 0
            flagd = False
    pxder = 0
#print ("cod: " +str(cod))
"""
for i in range (0,20):
    print(rd[i])
"""
print("calculando promedios d")
h=0.0
for i in range (cod):
    h = rd[i] + h
pxder = h/float(cod)

sumd=0.0
for i in range(0,cod):
	sumd=sumd+((rd[i]-pxder)*(rd[i]-pxder))
sumd=sumd/float(cod-1)
sumd=math.sqrt(sumd)
print("sumd:"+str(sumd))
print("calculando promedios i")
h = 0.0
for i in range (0,coi):
    h = ri[i] + h
pxizq = h/float(coi)

sumi=0.0
for i in range(0,coi):
	sumi=sumi+((ri[i]-pxizq)*(ri[i]-pxizq))
sumi=sumi/float(coi-1)
sumi=math.sqrt(sumi)
print("sumi:"+str(sumi))


tiempo_final = time() 
wb=Workbook()
desktop = os.path.normpath(os.path.expanduser("~/Desktop"))
now = datetime.datetime.now()
date=(str(now))
date=date.replace(":","-")
date=date.replace(".","-")
#print("fecha editada"+str(date))
filepath=desktop+"/demo"+str(date)+".xlsx"
#print(filepath)
sheet=wb.active
sheet.cell(row=1, column=2).value ="Rz Izquierda"
sheet.cell(row=1, column=3).value ="Rz Derecha"
sheet.cell(row=1, column=4).value ="Rz promedio"
sheet.cell(row=2, column=2).value =pxizq
sheet.cell(row=2, column=3).value =pxder
prom=(pxizq+pxder)/2
sheet.cell(row=2, column=4).value =prom
wb.save(filepath)

print("tiempo ejecucion fotografia 6000x4000:"+str(tiempo_final-tiempo_inicial))
print("promediod:"+str(pxder))
print("promedioi:"+str(pxizq))
print("promedio fotografia: "+str(prom))